import requests
from bs4 import BeautifulSoup
import time
import pandas as pd
import openpyxl
from selenium import webdriver

P_range = input("How many pages do you want to scrape: ")
sheet_num = input("Which sheet: ")

P_range_int = int(P_range)
P_range_int = P_range_int + 1

# Define the user-agent and headers
user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36'
headers = {'User-Agent': user_agent, 'Accept-Language': 'en-US,en;q=0.5'}

# Define the base URL to scrape
base_url = 'https://www.foreclosure.com/listings/montgomery-county-tn/preforeclosures/?q=Montgomery%20County,%20TN&pg={}&loc=Montgomery%20County,%20TN&view=list&'

# Define the delay between requests (in seconds)
delay = 0

# Load the existing Excel file
workbook = openpyxl.load_workbook('C:/Users/example/OneDrive/Desktop/datascraper.xlsx')#I replaced my account name with example for the obvious reason of privacy.

# Select the sheet you want to append data to
sheet = workbook[sheet_num]

# Determine the next available row in the sheet
next_row = sheet.max_row + 1

# Loop over all pages
for page in range(1, P_range_int):
    # Define the URL for the current page
    url = base_url.format(page)

    # Make the request
    response = requests.get(url, headers=headers)

    # Wait for the specified delay before making another request
    time.sleep(delay)

    # Parse the HTML content of the response using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find all the h1 and h2 tags on the page
    html1_tags = soup.select('a.address')
    html2_tags = soup.select('div.bedbathsizetype')
    html3_tags = soup.select('span.tdprice')

    # Append the data to the sheet
    for i, tag in enumerate(html1_tags):
        sheet.cell(row=next_row+i, column=1, value=tag.get_text())

    for i, tag in enumerate(html2_tags):
        sheet.cell(row=next_row+i, column=7, value=tag.get_text())

    for i, tag in enumerate(html3_tags):
        sheet.cell(row=next_row+i, column=15, value=tag.get_text())

    # Update the next available row
    next_row += len(html1_tags)

# Save the changes to the Excel file
workbook.save('C:/Users/example/OneDrive/Desktop/datascraper.xlsx')